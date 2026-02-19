# src/main.py
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np

from PySide6.QtCore import Qt
from PySide6.QtGui import QPixmap
from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QTabWidget,
    QGroupBox,
    QLabel,
    QLineEdit,
    QPushButton,
    QMessageBox,
    QFormLayout,
    QScrollArea,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QFileDialog,
    QSlider,
    QTextBrowser,
)

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle

from core import Params, StrokeCfg, simulate


# ----------------------------
# Helpers
# ----------------------------
def to_float(le: QLineEdit, name: str) -> float:
    try:
        return float(le.text().strip())
    except Exception:
        raise ValueError(f"{name} sayısal olmalı.")


def to_int(le: QLineEdit, name: str) -> int:
    try:
        return int(float(le.text().strip()))
    except Exception:
        raise ValueError(f"{name} tam sayı olmalı.")


def resource_path(relative: str) -> Path:
    # one-file bundle
    if hasattr(sys, "_MEIPASS"):
        return Path(getattr(sys, "_MEIPASS")) / relative

    # one-dir bundle: next to exe
    exe_dir = Path(sys.executable).resolve().parent
    p1 = exe_dir / relative
    if p1.exists():
        return p1

    # dev: project root assumed one above src/
    proj = Path(__file__).resolve().parents[1]
    return proj / relative


class MplCanvas(FigureCanvas):
    def __init__(self):
        fig = Figure()
        super().__init__(fig)
        self.fig = fig


# ----------------------------
# Image label (auto fit, no scroll)
# ----------------------------
class AutoFitImage(QLabel):
    def __init__(self):
        super().__init__()
        self._pix_src: QPixmap | None = None
        self.setAlignment(Qt.AlignCenter)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_source_pixmap(self, pix: QPixmap):
        self._pix_src = pix
        self._apply_scaled()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._apply_scaled()

    def _apply_scaled(self):
        if self._pix_src is None:
            return
        w = max(10, self.width() - 10)
        h = max(10, self.height() - 10)
        pix2 = self._pix_src.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.setPixmap(pix2)


# ----------------------------
# Main UI
# ----------------------------
class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Mechanism Analyzer")
        self.resize(1280, 820)

        # last outputs
        self.out: dict | None = None

        # sim precompute
        self.sim_O = None      # (2,)
        self.sim_A = None      # (Nt,2)
        self.sim_B = None      # (Nt,2)
        self.sim_xlim = None
        self.sim_ylim = None

        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        # Banner (logo + title)
        root.addLayout(self._build_banner())

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        # Tabs
        self.tab_input = QWidget()
        self.tab_sim = QWidget()
        self.tab_summary = QWidget()
        self.tab_angle = QWidget()
        self.tab_w = QWidget()
        self.tab_alpha = QWidget()
        self.tab_v = QWidget()
        self.tab_a = QWidget()
        self.tab_data = QWidget()
        self.tab_eq = QWidget()

        self.tabs.addTab(self.tab_input, "Giriş")
        self.tabs.addTab(self.tab_sim, "Simülasyon")
        self.tabs.addTab(self.tab_summary, "Özet")
        self.tabs.addTab(self.tab_angle, "Bıçak Açısı")
        self.tabs.addTab(self.tab_w, "Açısal Hız")
        self.tabs.addTab(self.tab_alpha, "Açısal İvme")
        self.tabs.addTab(self.tab_v, "Bıçak Hızı")
        self.tabs.addTab(self.tab_a, "Bıçak İvmesi")
        self.tabs.addTab(self.tab_data, "Veriler")
        self.tabs.addTab(self.tab_eq, "Denklemler")

        # disable result tabs until first run (everything except giriş)
        for i in range(1, self.tabs.count()):
            self.tabs.setTabEnabled(i, False)

        self._build_tab_input()
        self._build_tab_summary()
        self._build_tab_plots()
        self._build_tab_data()
        self._build_tab_sim()
        self._build_tab_equations()

    # ----------------------------
    # Banner
    # ----------------------------
    def _build_banner(self):
        lay = QHBoxLayout()
        lay.setSpacing(10)

        self.logo = QLabel()
        self.logo.setFixedHeight(46)
        self.logo.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)

        logo_path = resource_path("assets/roketsan_logo.png")
        if logo_path.exists():
            pix = QPixmap(str(logo_path))
            pix2 = pix.scaledToHeight(40, Qt.SmoothTransformation)
            self.logo.setPixmap(pix2)
        else:
            self.logo.setText("roketsan")

        title = QLabel("Mechanism Kinematic Analyzer (Teknik Demo)")
        title.setStyleSheet("font-weight:600; font-size: 14pt; color:#005F2C;")

        lay.addWidget(self.logo, 0)
        lay.addWidget(title, 0)
        lay.addStretch(1)
        return lay

    # ----------------------------
    # Tab 1: Input + Image (NO SCROLL on image)
    # ----------------------------
    def _build_tab_input(self):
        layout = QHBoxLayout(self.tab_input)
        layout.setSpacing(10)

        # Left: compact inputs (scrollable to avoid overlap)
        left = QVBoxLayout()
        left.setSpacing(8)

        box_geo = QGroupBox("Geometri")
        form_geo = QFormLayout(box_geo)
        form_geo.setVerticalSpacing(6)
        form_geo.setHorizontalSpacing(10)

        self.le_L1 = QLineEdit("28")
        self.le_L2 = QLineEdit("30.017")
        self.le_Ltip = QLineEdit("565")  # used for outputs (plots), not for sim view
        form_geo.addRow("Kol-1 Uzunluğu (A-B) [mm]", self.le_L1)
        form_geo.addRow("Kol-2 Uzunluğu (O-B) [mm]", self.le_L2)
        form_geo.addRow("Bıçak Uzunluğu [mm]", self.le_Ltip)
        note = QLabel("Not: Bıçak uzunluğu O noktasından itibaren ölçülür.")
        note.setWordWrap(True)
        note.setStyleSheet("color:#2D2D2D; font-weight:400;")
        form_geo.addRow(note)

        box_off = QGroupBox("Ofsetler")
        form_off = QFormLayout(box_off)
        form_off.setVerticalSpacing(6)
        form_off.setHorizontalSpacing(10)

        self.le_D1 = QLineEdit("9.042")
        self.le_D2 = QLineEdit("30")
        self.le_H1 = QLineEdit("26.5")
        self.le_H2 = QLineEdit("1")
        form_off.addRow("A X-Ofset (D1) [mm]", self.le_D1)
        form_off.addRow("O X-Ofset (D2) [mm]", self.le_D2)
        form_off.addRow("H1 [mm]", self.le_H1)
        form_off.addRow("H2 [mm]", self.le_H2)

        box_motion = QGroupBox("Hareket Tanımı")
        form_m = QFormLayout(box_motion)
        form_m.setVerticalSpacing(6)
        form_m.setHorizontalSpacing(10)

        self.le_t_end = QLineEdit("0.160")
        self.le_Nt = QLineEdit("3000")
        self.le_stroke = QLineEdit("50")
        form_m.addRow("Hareket Süresi [s]", self.le_t_end)
        form_m.addRow("Zaman Adım Sayısı", self.le_Nt)
        form_m.addRow("Stroke [mm]", self.le_stroke)

        btn_row = QHBoxLayout()

        self.btn_run = QPushButton("Hesapla")
        self.btn_reset = QPushButton("Reset")
        self.btn_run.clicked.connect(self.on_run)
        self.btn_reset.clicked.connect(self.on_reset)

        # ensure these are always green-styled via theme
        self.btn_run.setProperty("primaryAction", True)
        self.btn_reset.setProperty("primaryAction", True)

        btn_row.addWidget(self.btn_run)
        btn_row.addWidget(self.btn_reset)

        self.lbl_status = QLabel("Hazır.")
        self.lbl_status.setWordWrap(True)

        left.addWidget(box_geo)
        left.addWidget(box_off)
        left.addWidget(box_motion)
        left.addLayout(btn_row)
        left.addWidget(self.lbl_status)
        left.addStretch(1)

        left_container = QWidget()
        left_container.setLayout(left)
        left_container.setMaximumWidth(440)

        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setWidget(left_container)
        left_scroll.setMaximumWidth(470)

        # Right: mechanism image (NO scroll; auto-fit)
        right = QVBoxLayout()
        title = QLabel("Mekanizma Görseli")
        title.setStyleSheet("font-weight:600; color:#005F2C;")
        right.addWidget(title, 0, Qt.AlignLeft)

        self.img_label = AutoFitImage()
        right.addWidget(self.img_label, 1)

        layout.addWidget(left_scroll, 0)
        layout.addLayout(right, 1)

        self._load_mech_image()

    def _load_mech_image(self):
        img_path = resource_path("assets/mechanism.png")
        if img_path.exists():
            pix = QPixmap(str(img_path))
            self.img_label.set_source_pixmap(pix)
        else:
            self.img_label.setText("assets/mechanism.png bulunamadı.")

    def on_reset(self):
        self.le_L1.setText("28")
        self.le_L2.setText("30.017")
        self.le_Ltip.setText("565")
        self.le_D1.setText("9.042")
        self.le_D2.setText("30")
        self.le_H1.setText("26.5")
        self.le_H2.setText("1")
        self.le_t_end.setText("0.160")
        self.le_Nt.setText("3000")
        self.le_stroke.setText("50")
        self.lbl_status.setText("Hazır.")
        self.tabs.setCurrentIndex(0)

    # ----------------------------
    # Tab: Summary
    # ----------------------------
    def _build_tab_summary(self):
        lay = QVBoxLayout(self.tab_summary)
        self.lbl_summary = QLabel("Henüz sonuç yok. Giriş sekmesinden Hesapla.")
        self.lbl_summary.setWordWrap(True)
        lay.addWidget(self.lbl_summary)
        lay.addStretch(1)

    # ----------------------------
    # Plot tabs
    # ----------------------------
    def _build_tab_plots(self):
        self.canvas_angle = MplCanvas()
        self.ax_angle = self.canvas_angle.fig.add_subplot(111)
        QVBoxLayout(self.tab_angle).addWidget(self.canvas_angle)

        self.canvas_w = MplCanvas()
        self.ax_w = self.canvas_w.fig.add_subplot(111)
        QVBoxLayout(self.tab_w).addWidget(self.canvas_w)

        self.canvas_alpha = MplCanvas()
        self.ax_alpha = self.canvas_alpha.fig.add_subplot(111)
        QVBoxLayout(self.tab_alpha).addWidget(self.canvas_alpha)

        self.canvas_v = MplCanvas()
        self.ax_v = self.canvas_v.fig.add_subplot(111)
        QVBoxLayout(self.tab_v).addWidget(self.canvas_v)

        self.canvas_a = MplCanvas()
        self.ax_a = self.canvas_a.fig.add_subplot(111)
        QVBoxLayout(self.tab_a).addWidget(self.canvas_a)

    # ----------------------------
    # Tab: Data + export
    # ----------------------------
    def _build_tab_data(self):
        lay = QVBoxLayout(self.tab_data)

        top = QHBoxLayout()
        self.btn_export = QPushButton("Excel'e Aktar (.xlsx)")
        self.btn_export.clicked.connect(self.export_xlsx)
        top.addWidget(self.btn_export)
        top.addStretch(1)

        self.table = QTableWidget()
        lay.addLayout(top)
        lay.addWidget(self.table, 1)

    def export_xlsx(self):
        if self.out is None:
            QMessageBox.information(self, "Bilgi", "Önce hesaplama yapın.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            path, _ = QFileDialog.getSaveFileName(
                self, "Excel Kaydet", "mechanism_results.xlsx", "Excel Files (*.xlsx)"
            )
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            headers = self._table_headers()
            ws.append(headers)

            rows = self._table_rows()
            for r in rows:
                ws.append(r)

            for c in range(1, len(headers) + 1):
                col = get_column_letter(c)
                ws.column_dimensions[col].width = max(14, min(34, len(headers[c - 1]) + 2))

            wb.save(path)
            QMessageBox.information(self, "Tamam", "Excel dosyası kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel export başarısız: {e}")

    # ----------------------------
    # Tab: Simulation (FIXED: physical coordinates & link-length consistency)
    # ----------------------------
    def _build_tab_sim(self):
        lay = QVBoxLayout(self.tab_sim)

        top = QHBoxLayout()
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setMinimum(0)
        self.slider.setMaximum(0)
        self.slider.valueChanged.connect(self._sim_update)
        self.sim_info = QLabel("Hesaplama sonrası aktif olur.")
        self.sim_info.setStyleSheet("color:#1A1A1A;")
        top.addWidget(QLabel("Konum:"))
        top.addWidget(self.slider, 1)
        top.addWidget(self.sim_info, 0)
        lay.addLayout(top)

        self.canvas_sim = MplCanvas()
        self.ax_sim = self.canvas_sim.fig.add_subplot(111)
        lay.addWidget(self.canvas_sim, 1)

        self.ax_sim.set_aspect("equal", adjustable="box")
        self.ax_sim.grid(True, alpha=0.10)
        self.ax_sim.set_title("Mekanizma Simülasyonu (Slider + L1 + L2)")
        self.canvas_sim.fig.tight_layout()

        # Artists
        self.line_rail, = self.ax_sim.plot([], [], linewidth=2, linestyle="--", alpha=0.35)

        self.line_L1, = self.ax_sim.plot([], [], linewidth=4)  # A-B
        self.line_L2, = self.ax_sim.plot([], [], linewidth=4)  # O-B

        self.pt_O = self.ax_sim.scatter([], [], s=90, zorder=5)
        self.pt_A = self.ax_sim.scatter([], [], s=90, zorder=6)
        self.pt_B = self.ax_sim.scatter([], [], s=110, zorder=7)

        # Slider rectangle patch (centered at A)
        self.slider_rect = Rectangle((0, 0), 10.0, 6.0, linewidth=2, fill=True, zorder=4)
        self.ax_sim.add_patch(self.slider_rect)

        # Colors
        self.line_L1.set_color("#4D4D4D")
        self.line_L2.set_color("#4D4D4D")
        self.line_rail.set_color("#777777")

        self.pt_O.set_color("#1f77b4")  # blue
        self.pt_A.set_color("#ff7f0e")  # orange
        self.pt_B.set_color("#2ca02c")  # green

        self.slider_rect.set_edgecolor("#005F2C")
        self.slider_rect.set_facecolor("#CFE7DA")  # soft greenish

    def _sim_update(self, k: int):
        if self.out is None or self.sim_A is None or self.sim_B is None or self.sim_O is None:
            return
        k = int(k)

        O = self.sim_O
        A = self.sim_A[k]
        B = self.sim_B[k]

        # stable limits (no rescale jitter)
        if self.sim_xlim is not None and self.sim_ylim is not None:
            self.ax_sim.set_xlim(*self.sim_xlim)
            self.ax_sim.set_ylim(*self.sim_ylim)

        # rail y=0 across whole view
        if self.sim_xlim is not None:
            x0, x1 = self.sim_xlim
            self.line_rail.set_data([x0, x1], [0.0, 0.0])

        # Links
        self.line_L1.set_data([A[0], B[0]], [A[1], B[1]])
        self.line_L2.set_data([O[0], B[0]], [O[1], B[1]])

        # Points
        self.pt_O.set_offsets([O])
        self.pt_A.set_offsets([A])
        self.pt_B.set_offsets([B])

        # Slider rectangle centered at A
        w = float(self.slider_rect.get_width())
        h = float(self.slider_rect.get_height())
        self.slider_rect.set_xy((A[0] - w / 2.0, A[1] - h / 2.0))

        # Info
        t = float(self.out["t"][k])
        s = float(self.out["S34"][k])
        th1 = float(np.rad2deg(self.out["theta1"][k]))
        th2 = float(np.rad2deg(self.out["theta2"][k]))

        # Optional sanity check display (non-intrusive)
        L1_now = float(np.hypot(B[0] - A[0], B[1] - A[1]))
        L2_now = float(np.hypot(B[0] - O[0], B[1] - O[1]))
        self.sim_info.setText(
            f"t={t:.4f}s | Stroke={s:.2f}mm | θ1={th1:.2f}° | θ2={th2:.2f}° | "
            f"|AB|={L1_now:.3f} | |OB|={L2_now:.3f}"
        )

        self.canvas_sim.draw()

    def _prepare_sim(self, out: dict, p: Params):
        """
        Physical reference (as confirmed):
          S_phys = S + D1
          B_tot  = H1 + H2
          A = (-S_phys, 0)
          O = (-D2, B_tot)
          B = (-S_phys + L1*cos(theta1), L1*sin(theta1))

        This guarantees |AB| = L1 by construction and keeps the drawing consistent
        with the declared physical model. |OB| = L2 should also hold if solution is consistent.
        """
        S = out["S34"]
        th1 = out["theta1"]
        Btot = p.H1 + p.H2

        O = np.array([-p.D2, Btot], dtype=float)

        # A = (-(S + D1), 0)
        Ax = -(S + p.D1)
        Ay = np.zeros_like(S)
        A = np.column_stack([Ax, Ay])

        # B = (Ax + L1*cos(th1), L1*sin(th1))
        Bx = Ax + p.L1 * np.cos(th1)
        By = p.L1 * np.sin(th1)
        B = np.column_stack([Bx, By])

        self.sim_O = O
        self.sim_A = A
        self.sim_B = B

        # Fixed limits for stable view (no zoom changes)
        all_pts = np.vstack([A, B, np.tile(O, (len(S), 1))])
        xmin, ymin = all_pts.min(axis=0)
        xmax, ymax = all_pts.max(axis=0)

        span = float(max(xmax - xmin, ymax - ymin))
        pad = 0.25 * max(1.0, span)

        self.sim_xlim = (xmin - pad, xmax + pad)
        self.sim_ylim = (min(-pad * 0.15, ymin - pad * 0.15), ymax + pad)

        self.ax_sim.set_xlim(*self.sim_xlim)
        self.ax_sim.set_ylim(*self.sim_ylim)

        # slider range
        self.slider.setMaximum(len(S) - 1)
        self.slider.setValue(0)
        self._sim_update(0)

    # ----------------------------
    # Tab: Equations (FINAL formatted; readable + fast)
    # ----------------------------
    def _build_tab_equations(self):
        lay = QVBoxLayout(self.tab_eq)

        box = QGroupBox("Denklemler")
        v = QVBoxLayout(box)

        tb = QTextBrowser()
        tb.setOpenExternalLinks(False)

        # Note: keep content same; improve layout/readability.
        html = """
        <div style="font-family:Segoe UI, Arial; font-size:12pt; line-height:1.55; padding:14px; color:#1A1A1A;">
          <div style="color:#005F2C; font-weight:800; font-size:15pt; margin-bottom:10px;">
            Denklemler Sekmesi
          </div>

          <div style="margin:10px 0 6px; font-weight:800; color:#005F2C;">Konum Denklemleri</div>
          <div style="padding:10px 12px; border:1px solid #3A3A3A; border-radius:8px; background:#F4F6F5;">
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              D2 − (S + D1) + L1·cos(θ1) + L2·cos(θ2) = 0
            </div>
            <div style="height:8px;"></div>
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              −(H1 + H2) + L1·sin(θ1) + L2·sin(θ2) = 0
            </div>
          </div>

          <div style="margin:14px 0 6px; font-weight:800; color:#005F2C;">Hız Denklemleri</div>
          <div style="padding:10px 12px; border:1px solid #3A3A3A; border-radius:8px; background:#F4F6F5;">
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              −L1·sin(θ1)·θ̇1 − L2·sin(θ2)·θ̇2 = Ṡ
            </div>
            <div style="height:8px;"></div>
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              L1·cos(θ1)·θ̇1 + L2·cos(θ2)·θ̇2 = 0
            </div>
          </div>

          <div style="margin:14px 0 6px; font-weight:800; color:#005F2C;">İvme Denklemleri</div>
          <div style="padding:10px 12px; border:1px solid #3A3A3A; border-radius:8px; background:#F4F6F5;">
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              −L1·sin(θ1)·θ̈1 − L2·sin(θ2)·θ̈2 = S̈ + L1·cos(θ1)·(θ̇1)² + L2·cos(θ2)·(θ̇2)²
            </div>
            <div style="height:8px;"></div>
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              L1·cos(θ1)·θ̈1 + L2·cos(θ2)·θ̈2 = L1·sin(θ1)·(θ̇1)² + L2·sin(θ2)·(θ̇2)²
            </div>
          </div>

          <div style="margin:14px 0 6px; font-weight:800; color:#005F2C;">Bıçak İlişkileri</div>
          <div style="padding:10px 12px; border:1px solid #3A3A3A; border-radius:8px; background:#F4F6F5;">
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              δ<sub>tip</sub> = arctan(H2 / D2)
            </div>
            <div style="height:8px;"></div>
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              θ<sub>bıçak</sub> = θ2 + δ<sub>tip</sub>
            </div>
            <div style="height:8px;"></div>
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              θ̇<sub>bıçak</sub> = θ̇2
            </div>
            <div style="height:8px;"></div>
            <div style="font-family:Consolas, 'Courier New', monospace; font-size:12pt;">
              θ̈<sub>bıçak</sub> = θ̈2
            </div>
          </div>

        </div>
        """
        tb.setHtml(html)

        v.addWidget(tb, 1)
        lay.addWidget(box, 1)

    # ----------------------------
    # Run simulation and update tabs
    # ----------------------------
    def on_run(self):
        try:
            self.lbl_status.setText("Hesaplanıyor...")
            self.btn_run.setEnabled(False)
            QApplication.processEvents()

            p = Params(
                L1=to_float(self.le_L1, "Kol-1 Uzunluğu"),
                L2=to_float(self.le_L2, "Kol-2 Uzunluğu"),
                L_tip=to_float(self.le_Ltip, "Bıçak Uzunluğu"),
                D1=to_float(self.le_D1, "D1"),
                D2=to_float(self.le_D2, "D2"),
                H1=to_float(self.le_H1, "H1"),
                H2=to_float(self.le_H2, "H2"),
            )

            cfg = StrokeCfg(
                t_end=to_float(self.le_t_end, "Hareket Süresi"),
                Nt=to_int(self.le_Nt, "Zaman Adım Sayısı"),
                stroke_mm=to_float(self.le_stroke, "Stroke"),
            )

            if cfg.t_end <= 0:
                raise ValueError("Hareket Süresi > 0 olmalı.")
            if cfg.Nt < 50:
                raise ValueError("Zaman Adım Sayısı en az 50 olmalı (öneri: 1000+).")
            if cfg.stroke_mm <= 0:
                raise ValueError("Stroke > 0 olmalı.")
            if p.L1 <= 0 or p.L2 <= 0 or p.L_tip <= 0:
                raise ValueError("Uzunluklar > 0 olmalı.")

            out = simulate(p, cfg)
            self.out = out

            ok_ratio = float(np.mean(out["ok"])) * 100.0
            s_end = float(out["S34"][-1])
            ang_end = float(np.rad2deg(out["theta_tip"][-1]))

            self.lbl_status.setText(
                f"Son: Stroke={s_end:.2f} mm | Bıçak Açısı={ang_end:.2f}° | başarı={ok_ratio:.1f}%"
            )

            for i in range(1, self.tabs.count()):
                self.tabs.setTabEnabled(i, True)

            self._update_summary(p, cfg, out)
            self._update_plots(out)
            self._update_table(out)
            self._prepare_sim(out, p)

            self.tabs.setCurrentIndex(2)

            if ok_ratio < 99.0:
                QMessageBox.warning(
                    self,
                    "Uyarı",
                    "Bazı adımlarda kök bulma başarısız. 'Veriler' sekmesindeki ok sütununu kontrol edin."
                )

        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))
            self.lbl_status.setText("Hata oluştu.")
        finally:
            self.btn_run.setEnabled(True)

    def _update_summary(self, p: Params, cfg: StrokeCfg, out: dict):
        S = out["S34"]
        ang = np.rad2deg(out["theta_tip"])
        w = out["dtheta_tip"]
        alpha = out["ddtheta_tip"]

        txt = []
        txt.append("Parametre Özeti")
        txt.append("----------------")
        txt.append(f"Kol-1 (A-B) = {p.L1:.6g} mm")
        txt.append(f"Kol-2 (O-B) = {p.L2:.6g} mm")
        txt.append(f"Bıçak Uzunluğu = {p.L_tip:.6g} mm (O noktasından itibaren)")
        txt.append(f"D1 = {p.D1:.6g} mm, D2 = {p.D2:.6g} mm")
        txt.append(f"H1 = {p.H1:.6g} mm, H2 = {p.H2:.6g} mm, (H1+H2) = {(p.H1+p.H2):.6g} mm")
        txt.append("")
        txt.append("Hareket Tanımı")
        txt.append("--------------")
        txt.append(f"Hareket Süresi = {cfg.t_end:.6g} s")
        txt.append(f"Zaman Adım Sayısı = {cfg.Nt}")
        txt.append(f"Stroke = {cfg.stroke_mm:.6g} mm")
        txt.append("")
        txt.append("Sonuç Özeti")
        txt.append("-----------")
        txt.append(f"Stroke (son) = {S[-1]:.6g} mm")
        txt.append(f"Bıçak Açısı (son) = {ang[-1]:.6g} °")
        txt.append(f"Bıçak Açısı min/max = {ang.min():.6g} / {ang.max():.6g} °")
        txt.append(f"Açısal Hız min/max = {w.min():.6g} / {w.max():.6g} °/s")
        txt.append(f"Açısal İvme min/max = {alpha.min():.6g} / {alpha.max():.6g} °/s²")
        txt.append(f"Çözüm başarı oranı = {np.mean(out['ok'])*100:.3f} %")

        self.lbl_summary.setText("\n".join(txt))

    def _update_plots(self, out: dict):
        S = out["S34"]
        ang = np.rad2deg(out["theta_tip"])
        w = out["dtheta_tip"]
        alpha = out["ddtheta_tip"]

        self.ax_angle.clear()
        self.ax_angle.plot(S, ang, linewidth=2)
        self.ax_angle.set_xlabel("Stroke (mm)")
        self.ax_angle.set_ylabel("Bıçak Açısı (°)")
        self.ax_angle.grid(True, alpha=0.20)
        self.ax_angle.set_title(f"Bıçak Açısı vs Stroke | Son: {ang[-1]:.2f}°")
        self.canvas_angle.fig.tight_layout()
        self.canvas_angle.draw()

        self.ax_w.clear()
        self.ax_w.plot(S, w, linewidth=2)
        self.ax_w.set_xlabel("Stroke (mm)")
        self.ax_w.set_ylabel("Açısal Hız (°/s)")
        self.ax_w.grid(True, alpha=0.20)
        self.ax_w.set_title("Açısal Hız vs Stroke")
        self.canvas_w.fig.tight_layout()
        self.canvas_w.draw()

        self.ax_alpha.clear()
        self.ax_alpha.plot(S, alpha, linewidth=2)
        self.ax_alpha.set_xlabel("Stroke (mm)")
        self.ax_alpha.set_ylabel("Açısal İvme (°/s²)")
        self.ax_alpha.grid(True, alpha=0.20)
        self.ax_alpha.set_title("Açısal İvme vs Stroke")
        self.canvas_alpha.fig.tight_layout()
        self.canvas_alpha.draw()

        self.ax_v.clear()
        self.ax_v.plot(S, out["v_tan"], linewidth=2)
        self.ax_v.set_xlabel("Stroke (mm)")
        self.ax_v.set_ylabel("Bıçak Hızı (mm/s)")
        self.ax_v.grid(True, alpha=0.20)
        self.ax_v.set_title("Bıçak Hızı vs Stroke")
        self.canvas_v.fig.tight_layout()
        self.canvas_v.draw()

        self.ax_a.clear()
        self.ax_a.plot(S, out["a_tan"], linewidth=2, label="Bıçak İvmesi")
        self.ax_a.plot(S, out["a_norm"], linewidth=1.5, linestyle="--", label="Merkezcil İvme")
        self.ax_a.set_xlabel("Stroke (mm)")
        self.ax_a.set_ylabel("İvme (mm/s²)")
        self.ax_a.grid(True, alpha=0.20)
        self.ax_a.set_title("Bıçak İvmesi vs Stroke")
        self.ax_a.legend(loc="best")
        self.canvas_a.fig.tight_layout()
        self.canvas_a.draw()

    def _table_headers(self):
        return [
            "Stroke (mm)",
            "Bıçak Açısı (deg)",
            "Açısal Hız (rad/s)",
            "Açısal İvme (rad/s^2)",
            "Bıçak Hızı (mm/s)",
            "Bıçak İvmesi (mm/s^2)",
            "Merkezcil İvme (mm/s^2)",
            "ok",
        ]

    def _table_rows(self):
        out = self.out
        S = out["S34"]
        ang = np.rad2deg(out["theta_tip"])
        w = out["dtheta_tip"]
        alpha = out["ddtheta_tip"]
        v_t = out["v_tan"]
        a_t = out["a_tan"]
        a_n = out["a_norm"]
        ok = out["ok"]

        rows = []
        for i in range(len(S)):
            rows.append([
                float(S[i]),
                float(ang[i]),
                float(w[i]),
                float(alpha[i]),
                float(v_t[i]),
                float(a_t[i]),
                float(a_n[i]),
                bool(ok[i]),
            ])
        return rows

    def _update_table(self, out: dict):
        headers = self._table_headers()
        rows = self._table_rows()

        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(rows))

        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                if c == 7:
                    text = "true" if v else "false"
                else:
                    text = f"{v:.6g}"
                item = QTableWidgetItem(text)
                if c == 7 and (not v):
                    item.setBackground(Qt.red)
                self.table.setItem(r, c, item)

        self.table.resizeColumnsToContents()

    # ----------------------------
    # Apply Roketsan Theme (NO pure white)
    # + Rev-3: buttons always green, separators darker
    # ----------------------------
    def apply_theme(self):
        # Soft palette (no #FFFFFF)
        bg = "#EEF2F0"
        card = "#EEF2F0"
        field = "#F4F6F5"

        # Rev-3: darker separators/strips (near-black)
        sep = "#2F2F2F"      # pane/group/separators
        field_border = "#5A5A5A"  # inputs/tables

        self.setStyleSheet(f"""
        QWidget {{
            background-color: {bg};
            color: #1A1A1A;
            font-size: 11pt;
        }}

        /* Darker pane border (separator strips) */
        QTabWidget::pane {{
            background: {bg};
            border: 1px solid {sep};
            border-radius: 6px;
        }}

        QTabBar::tab {{
            background: #005F2C;
            color: white;
            padding: 8px 12px;
            margin-right: 2px;
            border-top-left-radius: 6px;
            border-top-right-radius: 6px;
        }}
        QTabBar::tab:selected {{
            background: #00843D;
        }}

        /* Rev-3: Buttons always green (even disabled) */
        QPushButton {{
            background-color: #00843D;
            color: white;
            padding: 6px 10px;
            border-radius: 6px;
            font-weight: 700;
            border: 1px solid #004D24;
        }}
        QPushButton:hover {{
            background-color: #005F2C;
        }}
        QPushButton:pressed {{
            background-color: #004D24;
        }}
        QPushButton:disabled {{
            background-color: #00843D;   /* keep green */
            color: rgba(255,255,255,230);
            border: 1px solid #004D24;
        }}

        /* Darker group borders (separator strips) */
        QGroupBox {{
            background-color: {card};
            font-weight: 700;
            border: 1px solid {sep};
            margin-top: 8px;
            border-radius: 8px;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 3px 0 3px;
            color: #005F2C;
        }}

        /* Inputs */
        QLineEdit, QTableWidget, QTextBrowser {{
            background-color: {field};
            border: 1px solid {field_border};
            border-radius: 6px;
            padding: 4px 6px;
        }}

        QHeaderView::section {{
            background-color: {card};
            border: 1px solid {sep};
            padding: 6px;
        }}

        QScrollArea {{
            background-color: {bg};
            border: 0px;
        }}
        QScrollArea QWidget {{
            background-color: {bg};
        }}
        """)


def main():
    app = QApplication(sys.argv)
    w = App()
    w.apply_theme()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
